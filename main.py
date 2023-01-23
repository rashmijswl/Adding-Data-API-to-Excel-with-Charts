import openpyxl
import random

items = ["Pens", "pencil", "Eraser", "Sharpner", "Refil ink", "Markers", "Books"]
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "Items"
sheet.cell(row = 1, column = 2).value = "Price"
sheet.cell(row = 1, column = 3).value = "Quantity"

for r in range(2, 9, 1):
  sheet.cell(row = r, column = 1).value = items[r-2]
  sheet.cell(row = r, column = 2).value = random.randint(1.0,5.0)
  sheet.cell(row = r, column = 3).value = random.randint(100,500)
  

for r in range(1, 9, 1):
  print(sheet.cell(row = r, column = 1).value, "\t" ,sheet.cell(row = r, column = 2).value, "\t",sheet.cell(row = r, column = 3).value)
  print("\n")


refObj = openpyxl.chart.Reference(sheet, min_col=3, min_row=2, max_col=3,max_row=8)
seriesObj = openpyxl.chart.Series(refObj, title='My stationary Items')

#Barchart
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'BarChart'
chartObj.append(seriesObj)



#linechart
chartObj_line = openpyxl.chart.LineChart()
chartObj_line.title = 'LineChart'
chartObj_line.append(seriesObj)

#piechart
chartObj_pie = openpyxl.chart.PieChart()
chartObj_pie.title = 'PieChart'
chartObj_pie.append(seriesObj)

#scatterchart
chartObj_sc = openpyxl.chart.ScatterChart()
x = openpyxl.chart.Reference(sheet, min_col=3, min_row=2, max_col=3,max_row=8)
y = openpyxl.chart.Reference(sheet, min_col=3, min_row=2, max_col=3,max_row=8)
s = openpyxl.chart.Series(y, xvalues=x)
chartObj_sc.append(s)



sheet.add_chart(chartObj, 'D1')
sheet.add_chart(chartObj_line, 'D17')
sheet.add_chart(chartObj_sc, 'N1')
sheet.add_chart(chartObj_pie, 'N17')


wb.save('Assignment3Rashmi_charts.xlsx')

